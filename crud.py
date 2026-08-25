import uuid
import time
import json
from typing import List, Dict, Tuple, Any, Optional
from datetime import date, datetime, timedelta
from fastapi import HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import func, case, or_
import models
import schemas


# ===========================
# DATE PARSER
# ===========================

def parse_date(date_value):
    if not date_value:
        return None

    # Product dates read from PostgreSQL may already be date objects.
    if hasattr(date_value, "year") and hasattr(date_value, "month"):
        return date_value

    date_str = str(date_value).strip()

    formats = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%Y",
        "%m-%Y",
        "%b %Y",
        "%B %Y",
        "%d %b %Y",
        "%d %B %Y",
        "%d.%m.%Y",
        "%d/%m/%y",
        "%m/%y",
    ]

    for fmt in formats:
        try:
            parsed_date = datetime.strptime(date_str, fmt)

            if fmt in [
                "%m/%Y",
                "%m-%Y",
                "%b %Y",
                "%B %Y",
                "%m/%y",
            ]:
                parsed_date = parsed_date.replace(day=1)

            return parsed_date.date()

        except ValueError:
            continue

    cleaned = (
        date_str.upper()
        .replace("EXPIRY", "")
        .replace("EXP.", "")
        .replace("EXP", "")
        .replace("BEST BEFORE", "")
        .replace("BESTBY", "")
        .replace(":", "")
        .strip()
    )

    for fmt in ["%m/%Y", "%m-%Y", "%b %Y", "%B %Y"]:
        try:
            return datetime.strptime(cleaned, fmt).replace(day=1).date()
        except ValueError:
            continue

    raise ValueError(f"Unsupported date format: {date_str}")


# ===========================
# PRODUCT STATUS
# ===========================

def calculate_product_status(expiry_date):
    today = datetime.today().date()
    days_remaining = (expiry_date - today).days

    if days_remaining < 0:
        product_status = "Expired"
    elif days_remaining <= 30:
        product_status = "Expiring Soon"
    else:
        product_status = "Safe"

    return days_remaining, product_status


# ===========================
# GET PRODUCTS
# ===========================

def get_products(db: Session, user_id: int):
    return (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.is_deleted == False,
        )
        .all()
    )


def get_product(db: Session, product_id: int, user_id: int):
    return (
        db.query(models.Product)
        .filter(
            models.Product.id == product_id,
            models.Product.user_id == user_id,
            models.Product.is_deleted == False,
        )
        .first()
    )


# ===========================
# CREATE PRODUCT
# ===========================

def create_product(
    db: Session,
    product: schemas.ProductCreate,
    user_id: int,
):
    expiry = parse_date(product.expiry_date)
    manufacturing = (
        parse_date(product.manufacturing_date)
        if product.manufacturing_date
        else None
    )

    days_remaining, product_status = calculate_product_status(expiry)

    effective_price = product.price if (product.price is not None and product.price > 0) else product.unit_price
    calc_total_price = product.total_price if product.total_price > 0 else (product.quantity * effective_price)

    db_product = models.Product(
        user_id=user_id,
        product_name=product.product_name,
        brand=product.brand,
        category=product.category,
        batch_number=product.batch_number,
        quantity=product.quantity,
        unit_price=effective_price,
        total_price=calc_total_price,
        manufacturing_date=manufacturing,
        expiry_date=expiry,
        days_remaining=days_remaining,
        status=product_status,
        image_path=product.image_path,
        ocr_text=product.ocr_text,
    )

    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    invalidate_restock_cache(user_id)

    return db_product


# ===========================
# UPDATE PRODUCT
# ===========================

def update_product(
    db: Session,
    product_id: int,
    product: schemas.ProductCreate,
    user_id: int,
):
    db_product = get_product(db, product_id, user_id)

    if db_product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found for this shop.",
        )

    expiry = parse_date(product.expiry_date)
    manufacturing = (
        parse_date(product.manufacturing_date)
        if product.manufacturing_date
        else None
    )

    days_remaining, product_status = calculate_product_status(expiry)

    effective_price = product.price if (product.price is not None and product.price > 0) else product.unit_price

    db_product.product_name = product.product_name
    db_product.brand = product.brand
    db_product.category = product.category
    db_product.batch_number = product.batch_number
    db_product.quantity = product.quantity
    db_product.unit_price = effective_price
    db_product.total_price = product.total_price if product.total_price > 0 else (product.quantity * effective_price)
    db_product.manufacturing_date = manufacturing
    db_product.expiry_date = expiry
    db_product.days_remaining = days_remaining
    db_product.status = product_status
    db_product.image_path = product.image_path
    db_product.ocr_text = product.ocr_text

    db.commit()
    invalidate_restock_cache(user_id)
    return db_product


# ===========================
# DELETE PRODUCT
# ===========================

def delete_product(
    db: Session,
    product_id: int,
    user_id: int,
):
    db_product = get_product(db, product_id, user_id)

    if db_product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found for this shop.",
        )

    db_product.is_deleted = True
    db_product.deleted_at = datetime.utcnow()
    db_product.deleted_by = user_id
    db.commit()

    return {"message": "Product moved to Recently Deleted, recoverable for 60 days.", "is_deleted": True}


# ===========================
# SELL / PURCHASE TRANSACTION
# ===========================

def create_transaction(
    db: Session,
    transaction: schemas.TransactionCreate,
    shop_id: int,
):
    """
    Creates one purchase or sale transaction.
    Locks row to prevent overselling.
    """
    try:
        product = (
            db.query(models.Product)
            .filter(
                models.Product.id == transaction.product_id,
                models.Product.user_id == shop_id,
                models.Product.is_deleted == False,
            )
            .with_for_update()
            .first()
        )

        if product is None:
            raise HTTPException(
                status_code=404,
                detail="Product not found for this shop.",
            )

        subtotal = transaction.unit_price * transaction.quantity
        discount_amount = 0.0

        if transaction.transaction_type == "sell":
            if product.quantity < transaction.quantity:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"Insufficient stock. Available quantity: "
                        f"{product.quantity}."
                    ),
                )

            if transaction.discount_type == "flat":
                discount_amount = transaction.discount_value or 0.0

            elif transaction.discount_type == "percent":
                discount_amount = (
                    subtotal * (transaction.discount_value or 0.0) / 100
                )

            if discount_amount > subtotal:
                raise HTTPException(
                    status_code=422,
                    detail="Discount cannot be greater than the bill subtotal.",
                )

            final_price = subtotal - discount_amount
            product.quantity -= transaction.quantity

        else:
            final_price = subtotal
            product.quantity += transaction.quantity

        new_transaction = models.InventoryTransaction(
            transaction_id=(
                f"TXN-{datetime.utcnow():%Y%m%d%H%M%S}-"
                f"{uuid.uuid4().hex[:8].upper()}"
            ),
            shop_id=shop_id,
            product_id=product.id,
            transaction_type=transaction.transaction_type,
            quantity=transaction.quantity,
            unit_price=transaction.unit_price,
            discount_type=transaction.discount_type,
            discount_value=transaction.discount_value,
            final_price=final_price,
        )

        db.add(new_transaction)
        db.commit()
        db.refresh(new_transaction)

        return new_transaction

    except HTTPException:
        db.rollback()
        raise

    except Exception:
        db.rollback()
        raise


# ===========================
# POS / COUNTER SALES TRANSACTION
# ===========================

def generate_bill_number() -> str:
    return f"BILL-{uuid.uuid4().hex[:6].upper()}"


def create_sale_transaction(
    db: Session,
    sale_data: schemas.SaleCreate,
    user_id: int,
    current_user: Optional[models.User] = None,
):
    """
    Completes one pharmacy bill in a single optimized database transaction.
    """
    try:
        if current_user is not None:
            shop = current_user
        else:
            shop = (
                db.query(models.User)
                .filter(models.User.id == user_id)
                .first()
            )
            if shop is None:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="Shop user not found.",
                )

        # Batch-fetch and lock all distinct product rows in a single DB round-trip
        product_ids = list({item.product_id for item in sale_data.items})
        locked_products = (
            db.query(models.Product)
            .filter(
                models.Product.id.in_(product_ids),
                models.Product.user_id == user_id,
                models.Product.is_deleted == False,
            )
            .with_for_update()
            .all()
        )
        prod_map = {p.id: p for p in locked_products}

        prepared_items = []
        gross_subtotal = 0.0
        line_discount_total = 0.0
        default_shop_gst = shop.default_gst_percentage or 12.0

        # Process every cart item using batch-loaded products in memory
        for item in sale_data.items:
            product = prod_map.get(item.product_id)
            if product is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Product ID {item.product_id} was not found for this shop.",
                )

            is_strip = str(item.unit_type or "strip").lower() in ["strip", "pack"]

            # Decide the sale price.
            if item.unit_price is not None:
                unit_price = item.unit_price
            elif is_strip:
                unit_price = product.unit_price
            else:
                if product.price_per_unit is not None and product.price_per_unit > 0:
                    unit_price = product.price_per_unit
                elif product.loose_tablet_price is not None and product.loose_tablet_price > 0:
                    unit_price = product.loose_tablet_price
                elif product.units_per_pack is not None and product.units_per_pack > 0:
                    unit_price = product.unit_price / product.units_per_pack
                elif product.tablets_per_strip is not None and product.tablets_per_strip > 0:
                    unit_price = product.unit_price / product.tablets_per_strip
                else:
                    unit_price = round(product.unit_price / 10.0, 2)

            # Stock deduction for a sealed strip sale.
            if is_strip:
                if product.quantity < item.quantity:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=(
                            f"Insufficient strip stock for '{product.product_name}'. "
                            f"Available: {product.quantity}, requested: {item.quantity}."
                        ),
                    )

                product.quantity -= item.quantity

            # Stock deduction for loose tablet sale.
            else:
                tablets_per_pack = product.tablets_per_strip or product.units_per_pack or 10
                if product.tablets_per_strip is None or product.tablets_per_strip <= 0:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=(
                            f"tablets_per_strip is required to sell "
                            f"'{product.product_name}' as loose tablets."
                        ),
                    )

                available_loose_tablets = (
                    product.loose_tablet_stock
                    + (product.quantity * product.tablets_per_strip)
                )

                if available_loose_tablets < item.quantity:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=(
                            f"Insufficient loose-tablet stock for '{product.product_name}'. "
                            f"Available: {available_loose_tablets}, "
                            f"requested: {item.quantity}."
                        ),
                    )

                # Open the minimum number of sealed strips required.
                tablets_needed_from_sealed_stock = max(
                    0,
                    item.quantity - product.loose_tablet_stock,
                )

                if tablets_needed_from_sealed_stock > 0:
                    strips_to_open = (
                        tablets_needed_from_sealed_stock
                        + product.tablets_per_strip
                        - 1
                    ) // product.tablets_per_strip

                    product.quantity -= strips_to_open
                    product.loose_tablet_stock += (
                        strips_to_open * product.tablets_per_strip
                    )

                product.loose_tablet_stock -= item.quantity

            gross_line_total = round(unit_price * item.quantity, 2)
            line_discount = round(item.discount or 0.0, 2)

            if line_discount > gross_line_total:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"Line discount cannot exceed the price of "
                        f"'{product.product_name}'."
                    ),
                )

            taxable_line_total = round(gross_line_total - line_discount, 2)

            item_hsn = getattr(product, "hsn_code", None) or "3004"
            gst_percentage = (
                item.gst_percentage
                if item.gst_percentage is not None
                else (
                    product.gst_percentage
                    if (product.gst_percentage is not None and product.gst_percentage > 0)
                    else (
                        getattr(product, "gst_rate", None)
                        if (getattr(product, "gst_rate", None) is not None and product.gst_rate > 0)
                        else default_shop_gst
                    )
                )
            )

            gross_subtotal += gross_line_total
            line_discount_total += line_discount

            prepared_items.append(
                {
                    "product": product,
                    "item": item,
                    "unit_price": unit_price,
                    "gross_line_total": gross_line_total,
                    "line_discount": line_discount,
                    "taxable_line_total": taxable_line_total,
                    "gst_percentage": gst_percentage,
                }
            )

        # Feature 5: Customer Lookup & Auto-apply patient fixed discount (Single Combined Query)
        customer_record = None
        target_cust_name = (sale_data.customer_name or "").strip()
        is_walkin = (not target_cust_name or target_cust_name.lower() in ["walk-in customer", "walkin", "cash customer", ""])

        if sale_data.customer_phone and sale_data.customer_phone.strip() not in ["", "N/A"]:
            customer_record = db.query(models.Customer).filter(
                models.Customer.user_id == user_id,
                models.Customer.phone == sale_data.customer_phone.strip()
            ).first()
        elif not is_walkin:
            customer_record = db.query(models.Customer).filter(
                models.Customer.user_id == user_id,
                func.lower(models.Customer.name) == target_cust_name.lower()
            ).first()

        if customer_record and customer_record.fixed_discount_percent > 0 and (sale_data.discount_type is None or sale_data.discount_value == 0):
            sale_data.discount_type = "percent"
            sale_data.discount_value = customer_record.fixed_discount_percent

        taxable_subtotal = round(gross_subtotal - line_discount_total, 2)

        # Calculate bill-level discount.
        if sale_data.discount_type == "percent":
            bill_discount_amount = round(
                taxable_subtotal * sale_data.discount_value / 100,
                2,
            )
        else:
            bill_discount_amount = round(sale_data.discount_value, 2)

        if bill_discount_amount > taxable_subtotal:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Bill discount cannot exceed the taxable subtotal.",
            )

        bill_number = generate_bill_number()
        sale_items_to_create = []
        total_taxable_value = 0.0
        total_cgst = 0.0
        total_sgst = 0.0
        total_igst = 0.0
        total_gst_amount = 0.0
        final_total = 0.0
        allocated_bill_discount = 0.0

        # Allocate bill discount across lines, then calculate CGST/SGST/IGST per line.
        for index, prepared in enumerate(prepared_items):
            is_last_item = index == len(prepared_items) - 1

            if is_last_item:
                allocated_discount = round(
                    bill_discount_amount - allocated_bill_discount,
                    2,
                )
            elif taxable_subtotal > 0:
                allocated_discount = round(
                    bill_discount_amount
                    * prepared["taxable_line_total"]
                    / taxable_subtotal,
                    2,
                )
                allocated_bill_discount += allocated_discount
            else:
                allocated_discount = 0.0

            final_taxable_line_total = round(
                prepared["taxable_line_total"] - allocated_discount,
                2,
            )

            hsn = getattr(prepared["product"], "hsn_code", None) or "3004"
            gst_pct = float(prepared["gst_percentage"])

            if getattr(sale_data, "is_interstate", False):
                cgst_r, cgst_a = 0.0, 0.0
                sgst_r, sgst_a = 0.0, 0.0
                igst_r = gst_pct
                igst_a = round(final_taxable_line_total * (igst_r / 100.0), 2)
            else:
                cgst_r = round(gst_pct / 2.0, 2)
                sgst_r = round(gst_pct / 2.0, 2)
                cgst_a = round(final_taxable_line_total * (cgst_r / 100.0), 2)
                sgst_a = round(final_taxable_line_total * (sgst_r / 100.0), 2)
                igst_r, igst_a = 0.0, 0.0

            line_tax = round(cgst_a + sgst_a + igst_a, 2)
            final_line_total = round(final_taxable_line_total + line_tax, 2)

            total_taxable_value += final_taxable_line_total
            total_cgst += cgst_a
            total_sgst += sgst_a
            total_igst += igst_a
            total_gst_amount += line_tax
            final_total += final_line_total

            sale_items_to_create.append(
                models.SaleItem(
                    product_id=prepared["product"].id,
                    product_name=prepared["product"].product_name,
                    hsn_code=hsn,
                    quantity=prepared["item"].quantity,
                    unit_type=prepared["item"].unit_type,
                    unit_price=prepared["unit_price"],
                    discount=prepared["line_discount"] + allocated_discount,
                    gst_percentage=gst_pct,
                    gst_amount=line_tax,
                    taxable_value=final_taxable_line_total,
                    cgst_rate=cgst_r,
                    cgst_amount=cgst_a,
                    sgst_rate=sgst_r,
                    sgst_amount=sgst_a,
                    igst_rate=igst_r,
                    igst_amount=igst_a,
                    total_with_tax=final_line_total,
                    total_price=final_line_total,
                    line_total=final_line_total,
                    batch_number=(
                        prepared["item"].batch_number
                        or prepared["product"].batch_number
                    ),
                    tablets_per_strip=prepared["product"].tablets_per_strip,
                )
            )

        # Tax Summary Table grouped by GST Rate
        tax_summary_dict = {}
        for item in sale_items_to_create:
            rate_key = float(item.gst_percentage)
            if rate_key not in tax_summary_dict:
                tax_summary_dict[rate_key] = {
                    "gst_rate": rate_key,
                    "taxable_value": 0.0,
                    "cgst_amount": 0.0,
                    "sgst_amount": 0.0,
                    "igst_amount": 0.0,
                    "total_tax": 0.0,
                }
            tax_summary_dict[rate_key]["taxable_value"] = round(tax_summary_dict[rate_key]["taxable_value"] + item.taxable_value, 2)
            tax_summary_dict[rate_key]["cgst_amount"] = round(tax_summary_dict[rate_key]["cgst_amount"] + item.cgst_amount, 2)
            tax_summary_dict[rate_key]["sgst_amount"] = round(tax_summary_dict[rate_key]["sgst_amount"] + item.sgst_amount, 2)
            tax_summary_dict[rate_key]["igst_amount"] = round(tax_summary_dict[rate_key]["igst_amount"] + item.igst_amount, 2)
            tax_summary_dict[rate_key]["total_tax"] = round(tax_summary_dict[rate_key]["total_tax"] + item.cgst_amount + item.sgst_amount + item.igst_amount, 2)

        tax_summary_list = list(tax_summary_dict.values())

        # Handle Pending Payment status and Customer linking
        payment_mode_upper = (sale_data.payment_method or "CASH").strip().upper()
        is_pending = (payment_mode_upper == "PENDING")
        payment_status = "PENDING" if is_pending else "PAID"

        if not is_walkin and target_cust_name:
            if not customer_record:
                # Create customer on the fly
                customer_record = models.Customer(
                    user_id=user_id,
                    name=target_cust_name,
                    phone=sale_data.customer_phone.strip() if (sale_data.customer_phone and sale_data.customer_phone.strip()) else "N/A",
                    pending_amount=round(final_total, 2) if is_pending else 0.0
                )
                db.add(customer_record)
                db.flush()
            elif is_pending:
                customer_record.pending_amount = round((customer_record.pending_amount or 0.0) + final_total, 2)

        db_sale = models.Sale(
            user_id=user_id,
            bill_number=bill_number,
            subtotal=round(gross_subtotal, 2),
            discount_amount=round(
                line_discount_total + bill_discount_amount,
                2,
            ),
            tax_amount=round(total_gst_amount, 2),
            total_amount=round(final_total, 2),
            is_interstate=getattr(sale_data, "is_interstate", False),
            total_taxable_value=round(total_taxable_value, 2),
            total_cgst=round(total_cgst, 2),
            total_sgst=round(total_sgst, 2),
            total_igst=round(total_igst, 2),
            tax_summary_json=json.dumps(tax_summary_list),
            gst_number=getattr(shop, "gstin", None) or shop.gst_number or "07AABCE1234F1Z5",
            gst_percentage=shop.default_gst_percentage,
            discount_type=sale_data.discount_type,
            discount_value=sale_data.discount_value,
            payment_method=payment_mode_upper,
            payment_status=payment_status,
            customer_id=customer_record.id if customer_record else None,
            customer_name=target_cust_name or "Walk-in Customer",
            customer_phone=sale_data.customer_phone,
            doctor_name=sale_data.doctor_name,
            doctor_reg_no=sale_data.doctor_reg_no,
            notes=sale_data.notes,
            return_status="completed",
            is_completed_on_mobile=True,
            items=sale_items_to_create,
        )

        db.add(db_sale)
        db.commit()
        invalidate_restock_cache(user_id)
        db_sale.items = sale_items_to_create
        return db_sale

    except HTTPException:
        db.rollback()
        raise

    except Exception:
        db.rollback()
        raise
# ===========================
# SALE RETURNS
# ===========================

def create_sale_return(
    db: Session,
    sale_id: int,
    return_data: schemas.SaleReturnCreate,
    user_id: int,
):
    sale = (
        db.query(models.Sale)
        .filter(
            models.Sale.id == sale_id,
            models.Sale.user_id == user_id,
        )
        .with_for_update()
        .first()
    )

    if sale is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Sale bill not found for this shop.",
        )

    if sale.return_status == "returned":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This bill is already fully returned.",
        )

    total_refund = 0.0
    return_items = []

    try:
        for item in return_data.items:
            sale_item = (
                db.query(models.SaleItem)
                .filter(
                    models.SaleItem.id == item.sale_item_id,
                    models.SaleItem.sale_id == sale.id,
                )
                .with_for_update()
                .first()
            )

            if sale_item is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Sale item ID {item.sale_item_id} not found.",
                )

            already_returned_quantity = (
                db.query(func.coalesce(func.sum(models.SaleReturnItem.quantity), 0))
                .filter(models.SaleReturnItem.sale_item_id == sale_item.id)
                .scalar()
            )

            remaining_returnable = sale_item.quantity - already_returned_quantity

            if item.quantity <= 0:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Return quantity must be greater than 0.",
                )

            if item.quantity > remaining_returnable:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"Cannot return {item.quantity} units of "
                        f"{sale_item.product_name}. Returnable quantity: "
                        f"{remaining_returnable}."
                    ),
                )

            product = (
                db.query(models.Product)
                .filter(
                    models.Product.id == sale_item.product_id,
                    models.Product.user_id == user_id,
                )
                .with_for_update()
                .first()
            )

            if product is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Product for {sale_item.product_name} not found.",
                )

            if sale_item.unit_type == "loose_tablet":
                product.loose_tablet_stock += item.quantity
            else:
                product.quantity += item.quantity

            refund_amount = round(
                sale_item.total_price * item.quantity / sale_item.quantity,
                2,
            )

            total_refund += refund_amount

            return_items.append(
                models.SaleReturnItem(
                    sale_item_id=sale_item.id,
                    product_id=product.id,
                    quantity=item.quantity,
                    unit_price=sale_item.unit_price,
                    return_total=refund_amount,
                )
            )

        sale_return = models.SaleReturn(
            sale_id=sale.id,
            user_id=user_id,
            reason=return_data.reason,
            return_amount=round(total_refund, 2),
            items=return_items,
        )

        db.add(sale_return)

        total_sold_quantity = sum(item.quantity for item in sale.items)

        total_returned_quantity = (
            db.query(func.coalesce(func.sum(models.SaleReturnItem.quantity), 0))
            .join(models.SaleReturn)
            .filter(models.SaleReturn.sale_id == sale.id)
            .scalar()
        ) + sum(item.quantity for item in return_data.items)

        if total_returned_quantity >= total_sold_quantity:
            sale.return_status = "returned"
        else:
            sale.return_status = "partially_returned"

        db.commit()
        db.refresh(sale_return)

        return sale_return

    except HTTPException:
        db.rollback()
        raise

    except Exception:
        db.rollback()
        raise


def get_todays_returns(
    db: Session,
    user_id: int,
):
    today = datetime.utcnow().date()

    return (
        db.query(models.SaleReturn)
        .filter(
            models.SaleReturn.user_id == user_id,
            func.date(models.SaleReturn.created_at) == today,
        )
        .order_by(models.SaleReturn.created_at.desc())
        .all()
    )

# ===========================
# NOTIFICATION SETTINGS
# ===========================

def get_notification_settings(
    db: Session,
    user_id: int,
):
    settings = (
        db.query(models.NotificationSettings)
        .filter(models.NotificationSettings.user_id == user_id)
        .first()
    )

    if settings is None:
        settings = models.NotificationSettings(user_id=user_id)
        db.add(settings)
        db.commit()
        db.refresh(settings)

    return settings


def update_notification_settings(
    db: Session,
    user_id: int,
    data: schemas.NotificationSettingsCreate,
):
    settings = (
        db.query(models.NotificationSettings)
        .filter(models.NotificationSettings.user_id == user_id)
        .first()
    )

    if settings is None:
        settings = models.NotificationSettings(user_id=user_id)
        db.add(settings)

    settings.enabled = data.enabled
    settings.notify_before_days = data.notify_before_days
    settings.reminder_frequency = data.reminder_frequency
    settings.notification_time = data.notification_time
    settings.sound = data.sound
    settings.vibration = data.vibration

    db.commit()
    db.refresh(settings)

    return settings


# ===========================
# BULK IMPORT PRODUCTS
# ===========================

def import_products(
    db: Session,
    products: list,
    user_id: int,
):
    imported_products = []

    try:
        for product in products:
            expiry = parse_date(product.expiry_date)
            manufacturing = (
                parse_date(product.manufacturing_date)
                if product.manufacturing_date
                else None
            )

            days_remaining, product_status = calculate_product_status(expiry)

            db_product = models.Product(
                user_id=user_id,
                product_name=product.product_name,
                brand=getattr(product, "brand", ""),
                category=product.category,
                batch_number=product.batch_number,
                quantity=product.quantity,
                unit_price=getattr(product, "unit_price", 0),
                total_price=getattr(product, "total_price", 0),
                manufacturing_date=manufacturing,
                expiry_date=expiry,
                days_remaining=days_remaining,
                status=product_status,
                image_path="",
                ocr_text="Invoice Import",
                notified_expiring=False,
                notified_expired=False,
                last_notification_date=None,
            )

            db.add(db_product)
            imported_products.append(db_product)

        db.commit()

        for product in imported_products:
            db.refresh(product)

        return imported_products

    except Exception:
        db.rollback()
        raise


def bulk_update_gst(
    db: Session,
    user_id: int,
    hsn_code: str,
    gst_rate: float,
):
    products = (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.hsn_code == hsn_code,
        )
        .all()
    )
    for p in products:
        p.gst_rate = gst_rate
        p.gst_percentage = gst_rate
    db.commit()


def bulk_import_inventory(
    db: Session,
    user_id: int,
    cleaned_rows: List[Tuple[int, Dict[str, Any], List[str]]],
    on_duplicate: str = "skip",
) -> Dict[str, Any]:
    imported_names = list(set(data["product_name"].strip().lower() for _, data, _ in cleaned_rows))
    existing_products = (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.is_deleted == False,
            func.lower(models.Product.product_name).in_(imported_names)
        )
        .all()
    )
    existing_map = {p.product_name.strip().lower(): p for p in existing_products}

    rows_imported = 0
    rows_updated = 0
    rows_skipped = 0
    summary_warnings = []
    to_add = []

    for row_idx, data, warnings in cleaned_rows:
        for w in warnings:
            summary_warnings.append(
                {
                    "row": row_idx,
                    "product_name": data["product_name"],
                    "message": w,
                }
            )

        norm_name = data["product_name"].strip().lower()
        if norm_name in existing_map:
            if on_duplicate == "skip":
                rows_skipped += 1
                summary_warnings.append(
                    {
                        "row": row_idx,
                        "product_name": data["product_name"],
                        "message": f"Product '{data['product_name']}' already exists; skipped.",
                    }
                )
                continue
            elif on_duplicate in ["update", "overwrite"]:
                p = existing_map[norm_name]
                if on_duplicate == "update":
                    p.quantity += data["quantity"]
                else:
                    p.quantity = data["quantity"]

                p.unit_price = data["unit_price"]
                p.purchase_price = data["purchase_price"]
                p.hsn_code = data["hsn_code"]
                p.gst_rate = data["gst_rate"]
                p.gst_percentage = data["gst_rate"]
                if data["batch_number"]:
                    p.batch_number = data["batch_number"]
                if data["expiry_date"]:
                    p.expiry_date = data["expiry_date"]
                    p.days_remaining = data["days_remaining"]
                    p.status = data["status"]
                rows_updated += 1
        else:
            new_prod = models.Product(
                user_id=user_id,
                product_name=data["product_name"],
                unit_price=data["unit_price"],
                purchase_price=data["purchase_price"],
                hsn_code=data["hsn_code"],
                gst_rate=data["gst_rate"],
                gst_percentage=data["gst_rate"],
                quantity=data["quantity"],
                expiry_date=data["expiry_date"],
                batch_number=data["batch_number"],
                tablets_per_strip=data["tablets_per_strip"],
                category=data["category"],
                days_remaining=data["days_remaining"],
                status=data["status"],
            )
            to_add.append(new_prod)
            existing_map[norm_name] = new_prod
            rows_imported += 1

    if to_add:
        db.add_all(to_add)
    db.commit()
    return {
        "rows_imported": rows_imported,
        "rows_updated": rows_updated,
        "rows_skipped": rows_skipped,
        "warnings": summary_warnings,
    }


# ===========================
# CUSTOMER CRUD & AUTO-DISCOUNT
# ===========================

def get_customer_by_phone(db: Session, phone: str, user_id: int):
    return (
        db.query(models.Customer)
        .filter(
            models.Customer.user_id == user_id,
            models.Customer.phone == phone,
        )
        .first()
    )


def create_or_update_customer(db: Session, customer_data: schemas.CustomerCreate, user_id: int):
    existing = get_customer_by_phone(db, customer_data.phone, user_id)
    if existing:
        existing.name = customer_data.name
        if customer_data.email:
            existing.email = customer_data.email
        if customer_data.address:
            existing.address = customer_data.address
        existing.fixed_discount_percent = customer_data.fixed_discount_percent
        db.commit()
        db.refresh(existing)
        return existing

    new_cust = models.Customer(
        user_id=user_id,
        name=customer_data.name,
        phone=customer_data.phone,
        email=customer_data.email,
        address=customer_data.address,
        fixed_discount_percent=customer_data.fixed_discount_percent,
    )
    db.add(new_cust)
    db.commit()
    db.refresh(new_cust)
    return new_cust


def get_customers(db: Session, user_id: int):
    return db.query(models.Customer).filter(models.Customer.user_id == user_id).all()


# ===========================
# SALE RETURNS CRUD
# ===========================

def process_sale_return(db: Session, return_data: schemas.SaleReturnCreate, user_id: int):
    sale = (
        db.query(models.Sale)
        .filter(models.Sale.id == return_data.sale_id, models.Sale.user_id == user_id)
        .first()
    )
    if not sale:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Sale bill not found.",
        )

    if getattr(sale, "return_status", None) == "returned":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This bill is already fully returned.",
        )

    total_return_amount = 0.0
    return_items = []

    for item_req in return_data.items:
        sale_item = (
            db.query(models.SaleItem)
            .filter(
                models.SaleItem.id == item_req.sale_item_id,
                models.SaleItem.sale_id == sale.id,
            )
            .first()
        )
        if not sale_item:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Sale item ID {item_req.sale_item_id} not found in bill.",
            )

        if item_req.quantity > sale_item.quantity:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Cannot return {item_req.quantity} of '{sale_item.product_name}'. Maximum sold: {sale_item.quantity}.",
            )

        # Calculate prorated refund amount per unit using total_price safely
        item_total = getattr(sale_item, "total_price", None)
        if item_total is None:
            item_total = getattr(sale_item, "total_amount", None) or (sale_item.unit_price * sale_item.quantity)
        
        unit_refund = (item_total / sale_item.quantity) if sale_item.quantity > 0 else sale_item.unit_price
        item_return_total = round(unit_refund * item_req.quantity, 2)
        total_return_amount += item_return_total

        # Restore inventory stock
        product = db.query(models.Product).filter(models.Product.id == sale_item.product_id).first()
        if product:
            if getattr(sale_item, "unit_type", "strip") == "strip":
                product.quantity = (product.quantity or 0) + item_req.quantity
            else:
                product.loose_tablet_stock = (product.loose_tablet_stock or 0) + item_req.quantity

        return_item = models.SaleReturnItem(
            sale_item_id=sale_item.id,
            product_id=sale_item.product_id,
            quantity=item_req.quantity,
            unit_price=round(unit_refund, 2),
            return_total=item_return_total,
        )
        return_items.append(return_item)

    sale_return = models.SaleReturn(
        sale_id=sale.id,
        user_id=user_id,
        reason=return_data.reason or "Patient Return",
        return_amount=round(total_return_amount, 2),
        items=return_items,
    )
    db.add(sale_return)

    sale.return_status = "partially_returned"
    db.commit()
    db.refresh(sale_return)
    invalidate_restock_cache(user_id)
    return sale_return


def get_todays_returns(db: Session, user_id: int):
    today_start = datetime.combine(date.today(), datetime.min.time())
    returns = (
        db.query(models.SaleReturn)
        .filter(
            models.SaleReturn.user_id == user_id,
            models.SaleReturn.created_at >= today_start,
        )
        .order_by(models.SaleReturn.created_at.desc())
        .all()
    )
    return returns


# ===========================
# RANKED SEARCH & AUTOCOMPLETE
# ===========================

def search_products_ranked(db: Session, query: str, user_id: int, limit: int = 10):
    clean_q = query.strip().lower()
    if not clean_q:
        return []

    prefix_pattern = f"{clean_q}%"
    contains_pattern = f"%{clean_q}%"

    rank_case = case(
        (func.lower(models.Product.product_name).like(prefix_pattern), 1),
        else_=2
    )

    results = (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.product_name.ilike(contains_pattern)
        )
        .order_by(
            rank_case.asc(),
            models.Product.verified.desc(),
            models.Product.product_name.asc()
        )
        .limit(limit)
        .all()
    )
    return results


# ===========================
# GLOBAL MEDICINE CATALOG & STOCK PURCHASE
# ===========================

_CATALOG_CACHE = {}

def search_medicine_catalog(db: Session, query: str, limit: int = 20):
    clean_q = query.strip()
    if not clean_q:
        return []

    cache_key = (clean_q.lower(), limit)
    now = time.time()
    if cache_key in _CATALOG_CACHE:
        cached_res, ts = _CATALOG_CACHE[cache_key]
        if now - ts < 120:
            return cached_res

    prefix_pat = f"{clean_q}%"
    contains_pat = f"%{clean_q}%"

    # Stage 1: Fast prefix search on product_name
    prefix_results = (
        db.query(models.MedicineCatalog)
        .filter(models.MedicineCatalog.product_name.ilike(prefix_pat))
        .order_by(models.MedicineCatalog.verified.desc(), models.MedicineCatalog.product_name.asc())
        .limit(limit)
        .all()
    )

    if len(prefix_results) >= limit:
        _CATALOG_CACHE[cache_key] = (prefix_results, now)
        return prefix_results

    seen_ids = set(m.id for m in prefix_results)
    remaining = limit - len(prefix_results)

    # Stage 2: Prefix search on brand or composition (salt)
    brand_comp_prefix = (
        db.query(models.MedicineCatalog)
        .filter(
            ~models.MedicineCatalog.id.in_(seen_ids),
            (models.MedicineCatalog.brand.ilike(prefix_pat)) | (models.MedicineCatalog.composition.ilike(prefix_pat))
        )
        .order_by(models.MedicineCatalog.verified.desc(), models.MedicineCatalog.product_name.asc())
        .limit(remaining)
        .all()
    )

    results = prefix_results + brand_comp_prefix
    if len(results) >= limit:
        _CATALOG_CACHE[cache_key] = (results, now)
        return results

    for m in brand_comp_prefix:
        seen_ids.add(m.id)
    remaining = limit - len(results)

    # Stage 3: Substring / contains match on product_name, brand, or composition (salt)
    contains_results = (
        db.query(models.MedicineCatalog)
        .filter(
            ~models.MedicineCatalog.id.in_(seen_ids),
            (models.MedicineCatalog.product_name.ilike(contains_pat)) |
            (models.MedicineCatalog.brand.ilike(contains_pat)) |
            (models.MedicineCatalog.composition.ilike(contains_pat))
        )
        .order_by(models.MedicineCatalog.verified.desc(), models.MedicineCatalog.product_name.asc())
        .limit(remaining)
        .all()
    )

    final_results = results + contains_results
    _CATALOG_CACHE[cache_key] = (final_results, now)
    return final_results


def check_duplicate_batch(db: Session, user_id: int, product_name: str, batch_number: str):
    clean_name = product_name.strip().lower()
    clean_batch = batch_number.strip()
    
    existing = (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.is_deleted == False,
            func.lower(models.Product.product_name) == clean_name,
            models.Product.batch_number == clean_batch
        )
        .first()
    )
    if existing:
        return {
            "is_duplicate": True,
            "product_id": existing.id,
            "product_name": existing.product_name,
            "batch_number": existing.batch_number,
            "existing_quantity": existing.quantity,
            "expiry_date": existing.expiry_date.strftime("%Y-%m-%d") if existing.expiry_date else None,
            "message": f"This batch '{clean_batch}' already exists in your inventory with {existing.quantity} units."
        }
    return {"is_duplicate": False, "message": "No duplicate batch found."}


def create_custom_medicine(db: Session, data: schemas.CustomMedicineCreate):
    clean_name = data.product_name.strip()
    
    # Check if catalog entry already exists
    existing = db.query(models.MedicineCatalog).filter(func.lower(models.MedicineCatalog.product_name) == clean_name.lower()).first()
    if existing:
        return existing

    catalog_entry = models.MedicineCatalog(
        product_name=clean_name,
        brand=data.brand,
        category=data.category or "allopathy",
        composition=data.composition,
        hsn_code=data.hsn_code or "3004",
        gst_rate=data.gst_rate or 12.0,
        default_price=data.default_price or 0.0,
        tablets_per_strip=data.tablets_per_strip or 10,
        units_per_pack=data.tablets_per_strip or 10,
        pack_size_label=data.pack_size_label or f"strip of {data.tablets_per_strip or 10} tablets",
        verified=True
    )
    db.add(catalog_entry)
    db.commit()
    db.refresh(catalog_entry)
    return catalog_entry


def add_real_inventory_item(db: Session, data: schemas.InventoryAddRequest, user_id: int, do_commit: bool = True):
    exp_date = parse_date(data.expiry_date)
    mfg_date = parse_date(data.manufacturing_date) if data.manufacturing_date else None

    if not exp_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid expiry date format. Use YYYY-MM-DD.",
        )

    days_remaining = (exp_date - date.today()).days

    if days_remaining < 0:
        product_status = "Expired"
    elif days_remaining <= 30:
        product_status = "Expiring Soon"
    else:
        product_status = "Safe"

    existing = None
    if data.duplicate_mode != "separate":
        existing = (
            db.query(models.Product)
            .filter(
                models.Product.user_id == user_id,
                models.Product.is_deleted == False,
                func.lower(models.Product.product_name) == data.product_name.strip().lower(),
                models.Product.batch_number == data.batch_number.strip(),
                models.Product.expiry_date == exp_date
            )
            .first()
        )

    units = data.units_per_pack or 10
    calc_per_unit = round(data.unit_price / units, 2) if (units and units > 0 and data.unit_price > 0) else None

    if existing and data.duplicate_mode != "separate":
        existing.quantity += data.quantity
        existing.purchase_price = data.purchase_price
        existing.unit_price = data.unit_price
        existing.units_per_pack = units
        existing.price_per_unit = calc_per_unit
        existing.loose_tablet_price = calc_per_unit
        existing.tablets_per_strip = units
        existing.total_price = existing.unit_price * existing.quantity
        existing.days_remaining = days_remaining
        existing.status = product_status
        if data.supplier_id:
            existing.supplier_id = data.supplier_id
        if data.document_id:
            existing.document_id = data.document_id
        if data.invoice_number:
            existing.invoice_number = data.invoice_number
        if do_commit:
            db.commit()
        return existing

    new_prod = models.Product(
        user_id=user_id,
        product_name=data.product_name.strip(),
        brand=data.brand,
        category=data.category or "allopathy",
        batch_number=data.batch_number.strip(),
        quantity=data.quantity,
        purchase_price=data.purchase_price,
        unit_price=data.unit_price,
        units_per_pack=units,
        price_per_unit=calc_per_unit,
        loose_tablet_price=calc_per_unit,
        tablets_per_strip=units,
        total_price=data.unit_price * data.quantity,
        hsn_code=data.hsn_code or "3004",
        gst_rate=data.gst_rate or 12.0,
        gst_percentage=data.gst_rate or 12.0,
        manufacturing_date=mfg_date,
        expiry_date=exp_date,
        days_remaining=days_remaining,
        status=product_status,
        supplier_id=data.supplier_id,
        document_id=data.document_id,
        invoice_number=data.invoice_number,
        verified=True,
        pack_size_verified=True,
        price_last_updated=datetime.utcnow()
    )

    db.add(new_prod)
    if do_commit:
        db.commit()
    return new_prod


# ===========================
# HSN TAX RATE CRUD & LOOKUP
# ===========================

DEFAULT_HSN_TAX_MAPPINGS = [
    {"hsn_code": "3004", "description": "Medicaments consisting of mixed or unmixed products for therapeutic/prophylactic uses", "gst_rate": 12.0, "category": "pharma", "is_life_saving": False},
    {"hsn_code": "3003", "description": "Medicaments (excluding goods of 3002, 3005 or 3006) for therapeutic/prophylactic uses", "gst_rate": 12.0, "category": "pharma", "is_life_saving": False},
    {"hsn_code": "3002", "description": "Vaccines, toxins, cultures of micro-organisms & specified life-saving drugs", "gst_rate": 5.0, "category": "pharma", "is_life_saving": True},
    {"hsn_code": "3001", "description": "Glands & organs for organo-therapeutic uses, heparin & extracts", "gst_rate": 5.0, "category": "pharma", "is_life_saving": False},
    {"hsn_code": "3005", "description": "Wadding, gauze, bandages, adhesive dressings & similar medical items", "gst_rate": 12.0, "category": "medical_devices", "is_life_saving": False},
    {"hsn_code": "3006", "description": "Pharmaceutical goods (sterile surgical catgut, blood-grouping reagents, ostomy appliances)", "gst_rate": 12.0, "category": "medical_devices", "is_life_saving": False},
    {"hsn_code": "3304", "description": "Beauty, skincare, medicated cosmetics or skin preparations", "gst_rate": 18.0, "category": "cosmetics", "is_life_saving": False},
    {"hsn_code": "2106", "description": "Food supplements, protein powders & dietary nutraceuticals", "gst_rate": 18.0, "category": "supplements", "is_life_saving": False},
    {"hsn_code": "9993", "description": "Healthcare services & medical equipment maintenance", "gst_rate": 18.0, "category": "services", "is_life_saving": False},
]

def seed_default_hsn_rates(db: Session):
    """Seed & update standard pharma HSN tax rates in reference table."""
    for item in DEFAULT_HSN_TAX_MAPPINGS:
        existing = db.query(models.HsnTaxRate).filter(models.HsnTaxRate.hsn_code == item["hsn_code"]).first()
        if not existing:
            db_item = models.HsnTaxRate(**item)
            db.add(db_item)
        else:
            existing.gst_rate = item["gst_rate"]
            existing.description = item["description"]
    db.commit()


def get_hsn_gst_rate(db: Session, hsn_code: Optional[str], product_name: Optional[str] = None, user_id: Optional[int] = None) -> Dict[str, Any]:
    """
    Looks up official GST rate by 4-digit or 6-digit HSN code.
    If unmapped or missing, logs to unmapped_hsn_logs for admin review.
    """
    seed_default_hsn_rates(db)

    clean_hsn = str(hsn_code).strip().replace(".", "") if hsn_code else ""
    
    if clean_hsn:
        # 1. Exact match
        match = db.query(models.HsnTaxRate).filter(models.HsnTaxRate.hsn_code == clean_hsn).first()
        if match:
            return {
                "hsn_code": match.hsn_code,
                "gst_rate": match.gst_rate,
                "description": match.description,
                "is_mapped": True,
                "is_life_saving": match.is_life_saving,
                "needs_manual_review": False
            }

        # 2. 4-Digit Prefix Match (e.g. 300410 -> 3004)
        if len(clean_hsn) > 4:
            prefix = clean_hsn[:4]
            prefix_match = db.query(models.HsnTaxRate).filter(models.HsnTaxRate.hsn_code == prefix).first()
            if prefix_match:
                return {
                    "hsn_code": clean_hsn,
                    "gst_rate": prefix_match.gst_rate,
                    "description": f"{prefix_match.description} (Matched 4-digit HSN prefix {prefix})",
                    "is_mapped": True,
                    "is_life_saving": prefix_match.is_life_saving,
                    "needs_manual_review": False
                }

    # 3. Unmapped or missing HSN handling -> Log for review
    if clean_hsn:
        log_entry = models.UnmappedHsnLog(
            hsn_code=clean_hsn,
            product_name=product_name,
            user_id=user_id,
            entered_gst_rate=None
        )
        db.add(log_entry)
        db.commit()

    return {
        "hsn_code": clean_hsn or "3004",
        "gst_rate": 5.0,  # GST 2.0 pharma baseline rate
        "description": f"Unmapped/Missing HSN ('{clean_hsn}') — Flagged for Shopkeeper Confirmation",
        "is_mapped": False,
        "is_life_saving": False,
        "needs_manual_review": True
    }


def get_all_hsn_rates(db: Session):
    seed_default_hsn_rates(db)
    return db.query(models.HsnTaxRate).order_by(models.HsnTaxRate.hsn_code.asc()).all()


def create_hsn_rate(db: Session, hsn_data: schemas.HsnTaxRateCreate):
    clean_hsn = hsn_data.hsn_code.strip().replace(".", "")
    existing = db.query(models.HsnTaxRate).filter(models.HsnTaxRate.hsn_code == clean_hsn).first()
    if existing:
        existing.gst_rate = hsn_data.gst_rate
        existing.description = hsn_data.description
        existing.category = hsn_data.category
        existing.is_life_saving = hsn_data.is_life_saving
        db.commit()
        db.refresh(existing)
        return existing

    new_hsn = models.HsnTaxRate(
        hsn_code=clean_hsn,
        description=hsn_data.description,
        gst_rate=hsn_data.gst_rate,
        category=hsn_data.category,
        is_life_saving=hsn_data.is_life_saving
    )
    db.add(new_hsn)
    db.commit()
    db.refresh(new_hsn)
    return new_hsn


def get_unmapped_hsn_logs(db: Session, limit: int = 50):
    return db.query(models.UnmappedHsnLog).order_by(models.UnmappedHsnLog.created_at.desc()).limit(limit).all()


# ===========================
# USER PROFILE & SETTINGS CRUD
# ===========================

def update_user_profile(db: Session, user_id: int, data: schemas.UserProfileUpdate):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User profile not found.")

    update_dict = data.model_dump(exclude_unset=True)
    for field, value in update_dict.items():
        if hasattr(user, field) and value is not None:
            setattr(user, field, value)

    db.commit()
    db.refresh(user)
    return user


def change_user_password(db: Session, user_id: int, req: schemas.PasswordChangeRequest):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")

    if not verify_password(req.current_password, user.password):
        raise HTTPException(status_code=400, detail="Current password is incorrect.")

    user.password = get_password_hash(req.new_password)
    db.commit()
    return {"message": "Password updated successfully."}


def update_sale_retrospective(
    db: Session,
    sale_id: int,
    update_data: schemas.SaleUpdateDesktop,
    user_id: int,
):
    """Update non-financial metadata on existing sale bill post-checkout."""
    sale = (
        db.query(models.Sale)
        .filter(models.Sale.id == sale_id, models.Sale.user_id == user_id)
        .first()
    )
    if not sale:
        raise HTTPException(status_code=404, detail="Sale bill not found or access denied.")

    if update_data.customer_name is not None:
        sale.customer_name = update_data.customer_name
    if update_data.customer_phone is not None:
        sale.customer_phone = update_data.customer_phone
    if update_data.payment_method is not None:
        sale.payment_method = update_data.payment_method
    if update_data.notes is not None:
        sale.notes = update_data.notes
    if update_data.doctor_name is not None:
        sale.doctor_name = update_data.doctor_name
    if update_data.doctor_reg_no is not None:
        sale.doctor_reg_no = update_data.doctor_reg_no

    db.commit()
    db.refresh(sale)
    return sale


def delete_sale_bill(db: Session, sale_id: int, user_id: int):
    """
    Deletes sale bill and safely restores unrefunded inventory stock,
    preventing double-restoration of already refunded items.
    """
    sale = (
        db.query(models.Sale)
        .filter(models.Sale.id == sale_id, models.Sale.user_id == user_id)
        .with_for_update()
        .first()
    )
    if not sale:
        raise HTTPException(status_code=404, detail="Bill not found or access denied.")

    # Restores stock for unreturned items only
    for item in sale.items:
        already_returned_qty = (
            db.query(func.coalesce(func.sum(models.SaleReturnItem.quantity), 0))
            .filter(models.SaleReturnItem.sale_item_id == item.id)
            .scalar()
        )
        net_restore_qty = max(0, item.quantity - already_returned_qty)

        if net_restore_qty > 0:
            product = (
                db.query(models.Product)
                .filter(models.Product.id == item.product_id, models.Product.user_id == user_id)
                .with_for_update()
                .first()
            )
            if product:
                if item.unit_type == "loose_tablet":
                    product.loose_tablet_stock += net_restore_qty
                else:
                    product.quantity += net_restore_qty

    db.delete(sale)
    db.commit()
    return {"message": f"Bill {sale.bill_number} deleted successfully and unrefunded inventory restored."}


# ==========================================
# SUPPLIER MANAGEMENT CRUD
# ==========================================

def create_supplier(db: Session, user_id: int, supplier_data: schemas.SupplierCreate):
    supplier = models.Supplier(
        user_id=user_id,
        name=supplier_data.name.strip(),
        contact_person=supplier_data.contact_person,
        phone=supplier_data.phone,
        email=supplier_data.email,
        address=supplier_data.address,
        gstin=supplier_data.gstin.strip() if supplier_data.gstin else None,
        state=supplier_data.state or "Delhi",
        payment_terms=supplier_data.payment_terms or "Net 30",
        notes=supplier_data.notes,
        status=supplier_data.status or "Active"
    )
    db.add(supplier)
    db.commit()
    db.refresh(supplier)
    return supplier


def get_suppliers(db: Session, user_id: int, query: Optional[str] = None, status: Optional[str] = None):
    q = db.query(models.Supplier).filter(models.Supplier.user_id == user_id)
    if status and status.upper() != "ALL":
        q = q.filter(models.Supplier.status == status)
    if query:
        clean_q = f"%{query.strip().lower()}%"
        q = q.filter(
            (func.lower(models.Supplier.name).like(clean_q)) |
            (func.lower(models.Supplier.gstin).like(clean_q)) |
            (func.lower(models.Supplier.phone).like(clean_q))
        )
    
    suppliers = q.order_by(models.Supplier.name.asc()).all()
    
    result = []
    for s in suppliers:
        docs = db.query(models.Document).filter(models.Document.supplier_id == s.id, models.Document.user_id == user_id).all()
        total_purchases = sum(d.total_amount for d in docs if d.total_amount)
        purchase_count = len(docs)
        last_doc = db.query(models.Document).filter(models.Document.supplier_id == s.id, models.Document.user_id == user_id).order_by(models.Document.created_at.desc()).first()
        last_date = last_doc.created_at.strftime("%Y-%m-%d") if last_doc else None

        res_dict = schemas.SupplierResponse.from_orm(s)
        res_dict.total_purchases = total_purchases
        res_dict.purchase_count = purchase_count
        res_dict.last_purchase_date = last_date
        result.append(res_dict)
        
    return result


def get_supplier_detail(db: Session, supplier_id: int, user_id: int):
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id, models.Supplier.user_id == user_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found.")
    return supplier


def update_supplier(db: Session, supplier_id: int, user_id: int, data: schemas.SupplierUpdate):
    supplier = get_supplier_detail(db, supplier_id, user_id)
    update_dict = data.model_dump(exclude_unset=True)
    for k, v in update_dict.items():
        if hasattr(supplier, k) and v is not None:
            setattr(supplier, k, v)
    db.commit()
    db.refresh(supplier)
    return supplier


def delete_supplier(db: Session, supplier_id: int, user_id: int):
    supplier = get_supplier_detail(db, supplier_id, user_id)
    supplier.status = "Inactive"
    db.commit()
    return {"message": f"Supplier '{supplier.name}' deactivated successfully."}


def get_supplier_purchases(db: Session, supplier_id: int, user_id: int):
    return db.query(models.Document).filter(models.Document.supplier_id == supplier_id, models.Document.user_id == user_id).order_by(models.Document.created_at.desc()).all()


def get_supplier_inventory(db: Session, supplier_id: int, user_id: int):
    return db.query(models.Product).filter(models.Product.supplier_id == supplier_id, models.Product.user_id == user_id).order_by(models.Product.product_name.asc()).all()


# ==========================================
# DOCUMENT MANAGEMENT CRUD
# ==========================================

def create_document(db: Session, user_id: int, title: str, doc_type: str, file_path: str, file_type: str, file_size: int, supplier_id: Optional[int] = None, invoice_number: Optional[str] = None, notes: Optional[str] = None):
    doc = models.Document(
        user_id=user_id,
        supplier_id=supplier_id,
        title=title,
        doc_type=doc_type,
        file_path=file_path,
        file_type=file_type,
        file_size=file_size,
        invoice_number=invoice_number,
        notes=notes,
        ocr_status="Processing"
    )
    db.add(doc)
    db.commit()
    return doc


def get_documents(db: Session, user_id: int, query: Optional[str] = None, doc_type: Optional[str] = None, status: Optional[str] = None, supplier_id: Optional[int] = None):
    q = db.query(models.Document).filter(models.Document.user_id == user_id)
    if doc_type and doc_type.lower() != "all":
        q = q.filter(models.Document.doc_type == doc_type)
    if status and status.lower() != "all":
        q = q.filter(models.Document.ocr_status == status)
    if supplier_id:
        q = q.filter(models.Document.supplier_id == supplier_id)
    if query:
        clean_q = f"%{query.strip().lower()}%"
        q = q.filter(
            (func.lower(models.Document.title).like(clean_q)) |
            (func.lower(models.Document.invoice_number).like(clean_q))
        )

    docs = q.order_by(models.Document.created_at.desc()).all()
    
    result = []
    for d in docs:
        supplier_name = d.supplier.name if d.supplier else None
        res_dict = schemas.DocumentResponse.from_orm(d)
        res_dict.supplier_name = supplier_name
        result.append(res_dict)
    return result


def get_document_detail(db: Session, document_id: int, user_id: int):
    doc = db.query(models.Document).filter(models.Document.id == document_id, models.Document.user_id == user_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found.")
    return doc


def confirm_document_and_update_stock(db: Session, document_id: int, user_id: int, confirm_req: schemas.DocumentConfirmRequest):
    doc = get_document_detail(db, document_id, user_id)
    
    if confirm_req.supplier_id:
        doc.supplier_id = confirm_req.supplier_id
    if confirm_req.invoice_number:
        doc.invoice_number = confirm_req.invoice_number
    if confirm_req.invoice_date:
        try:
            doc.invoice_date = datetime.strptime(confirm_req.invoice_date, "%Y-%m-%d").date()
        except Exception:
            pass
    if confirm_req.total_amount:
        doc.total_amount = confirm_req.total_amount

    doc.item_count = len(confirm_req.items)
    doc.ocr_status = "Verified"

    created_products = []
    for item in confirm_req.items:
        product_req = schemas.InventoryAddRequest(
            product_name=item.product_name,
            brand=item.brand,
            category=item.category or "allopathy",
            hsn_code=item.hsn_code or "3004",
            gst_rate=item.gst_rate or 12.0,
            batch_number=item.batch_number,
            quantity=item.quantity,
            purchase_price=item.purchase_price,
            unit_price=item.unit_price,
            expiry_date=item.expiry_date
        )
        prod = add_real_inventory_item(db, product_req, user_id, do_commit=False)
        
        # Link traceability
        prod.supplier_id = doc.supplier_id
        prod.document_id = doc.id
        prod.invoice_number = doc.invoice_number
        created_products.append(prod)

    db.commit()
    return {"message": f"Document verified successfully. {len(created_products)} item batches added to stock with full supplier traceability.", "document_id": doc.id}


def delete_document(db: Session, document_id: int, user_id: int):
    doc = get_document_detail(db, document_id, user_id)
    db.delete(doc)
    db.commit()
    return {"message": "Document deleted successfully."}


# ==========================================
# SOFT DELETE & 60-DAY RECOVERY CRUD
# ==========================================

def soft_delete_inventory_items(db: Session, stock_ids: List[int], user_id: int) -> Dict[str, Any]:
    """Soft-deletes selective inventory rows with 60-day recovery window."""
    now = datetime.utcnow()
    items = (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.id.in_(stock_ids),
            models.Product.is_deleted == False,
        )
        .all()
    )

    if not items:
        return {
            "success": True,
            "message": "0 items moved to Recently Deleted.",
            "deleted_count": 0,
            "stock_ids": []
        }

    affected_ids = []
    for item in items:
        item.is_deleted = True
        item.deleted_at = now
        item.deleted_by = user_id
        affected_ids.append(item.id)

    db.commit()
    return {
        "success": True,
        "message": f"{len(affected_ids)} items moved to Recently Deleted, recoverable for 60 days.",
        "deleted_count": len(affected_ids),
        "stock_ids": affected_ids
    }


def soft_delete_all_inventory_items(db: Session, user_id: int) -> Dict[str, Any]:
    """Soft-deletes all active (non-deleted) inventory rows for the authenticated shop."""
    now = datetime.utcnow()
    items = (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.is_deleted == False,
        )
        .all()
    )

    if not items:
        return {
            "success": True,
            "message": "No active stock items to delete.",
            "deleted_count": 0
        }

    count = len(items)
    for item in items:
        item.is_deleted = True
        item.deleted_at = now
        item.deleted_by = user_id

    db.commit()
    return {
        "success": True,
        "message": f"All {count} items moved to Recently Deleted, recoverable for 60 days.",
        "deleted_count": count
    }


def get_recently_deleted_inventory(db: Session, user_id: int) -> List[Dict[str, Any]]:
    """Returns soft-deleted items within the 60-day recovery window, sorted newest first."""
    cutoff = datetime.utcnow() - timedelta(days=60)
    items = (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.is_deleted == True,
            models.Product.deleted_at >= cutoff,
        )
        .order_by(models.Product.deleted_at.desc())
        .all()
    )

    result = []
    now = datetime.utcnow()
    for p in items:
        del_at = p.deleted_at or now
        days_passed = (now - del_at).days
        days_left = max(0, 60 - days_passed)
        result.append({
            "id": p.id,
            "user_id": p.user_id,
            "product_name": p.product_name,
            "brand": p.brand,
            "category": p.category,
            "batch_number": p.batch_number,
            "quantity": p.quantity,
            "unit_price": p.unit_price,
            "purchase_price": p.purchase_price,
            "expiry_date": p.expiry_date.strftime("%Y-%m-%d") if p.expiry_date else None,
            "days_remaining": p.days_remaining,
            "status": p.status,
            "is_deleted": p.is_deleted,
            "deleted_at": p.deleted_at,
            "deleted_by": p.deleted_by,
            "days_until_permanent_delete": days_left
        })

    return result


def restore_inventory_items(db: Session, stock_ids: List[int], user_id: int) -> Dict[str, Any]:
    """Restores soft-deleted items only if deleted within the 60-day recovery window."""
    cutoff = datetime.utcnow() - timedelta(days=60)
    items = (
        db.query(models.Product)
        .filter(
            models.Product.user_id == user_id,
            models.Product.id.in_(stock_ids),
            models.Product.is_deleted == True,
            models.Product.deleted_at >= cutoff,
        )
        .all()
    )

    if not items:
        return {
            "success": True,
            "message": "No recoverable items found for the provided IDs.",
            "restored_count": 0,
            "stock_ids": []
        }

    restored_ids = []
    for item in items:
        item.is_deleted = False
        item.deleted_at = None
        item.deleted_by = None
        restored_ids.append(item.id)

    db.commit()
    return {
        "success": True,
        "message": f"{len(restored_ids)} items restored to live inventory.",
        "restored_count": len(restored_ids),
        "stock_ids": restored_ids
    }


def purge_expired_soft_deleted_inventory(db: Session) -> int:
    """Permanently purges (hard-deletes) soft-deleted stock rows older than 60 days."""
    cutoff = datetime.utcnow() - timedelta(days=60)
    deleted_rows = (
        db.query(models.Product)
        .filter(
            models.Product.is_deleted == True,
            models.Product.deleted_at < cutoff,
        )
        .delete(synchronize_session=False)
    )
    db.commit()
    return deleted_rows


# ==========================================
# BULK EXCEL / CSV DIRECT INVENTORY IMPORT
# ==========================================

def parse_pack_units(pack_size_label: Optional[str]) -> int:
    if not pack_size_label:
        return 10
    import re
    s = str(pack_size_label).strip().lower()
    match_mult = re.search(r'(\d+)\s*[xX*]\s*(\d+)', s)
    if match_mult:
        return int(match_mult.group(1)) * int(match_mult.group(2))
    match_num = re.search(r'(\d+)', s)
    if match_num:
        val = int(match_num.group(1))
        if val > 0:
            return val
    return 10


def import_inventory_from_file(db: Session, user_id: int, file_bytes: bytes, filename: str) -> dict:
    """
    Directly parses an uploaded .xlsx, .xls, or .csv file (no Gemini AI required)
    and onboards medicine stock batches into the shop's active live inventory.
    """
    import io
    import csv
    import openpyxl

    rows_to_process = []
    fname_lower = filename.lower()

    if fname_lower.endswith((".xlsx", ".xls")):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
            all_rows = list(ws.iter_rows(values_only=True))

            if not all_rows:
                raise HTTPException(status_code=400, detail="Uploaded spreadsheet is empty.")

            header_idx = -1
            headers = []
            for r_idx, row in enumerate(all_rows):
                str_row = [str(c).strip().lower() for c in row if c is not None]
                if any("medicine" in c or "product" in c or "batch" in c for c in str_row):
                    header_idx = r_idx
                    headers = [str(c).strip().lower() if c is not None else "" for c in row]
                    break

            if header_idx == -1:
                raise HTTPException(
                    status_code=400,
                    detail="Could not find header row. Required columns: medicine_name, batch_no, expiry_date, quantity, mrp."
                )

            for row in all_rows[header_idx + 1:]:
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue
                row_dict = {}
                for h, val in zip(headers, row):
                    if h:
                        row_dict[h] = val
                rows_to_process.append(row_dict)

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read Excel spreadsheet: {str(e)}")

    elif fname_lower.endswith(".csv"):
        try:
            text_content = ""
            for enc in ["utf-8-sig", "utf-8", "latin1", "cp1252"]:
                try:
                    text_content = file_bytes.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue

            if not text_content:
                text_content = file_bytes.decode("utf-8", errors="replace")

            reader = csv.DictReader(io.StringIO(text_content))
            for row in reader:
                if any(v.strip() for v in row.values() if v):
                    clean_row = {k.strip().lower(): v for k, v in row.items() if k}
                    rows_to_process.append(clean_row)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV file: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Unsupported format. Please upload an .xlsx, .xls, or .csv file.")

    if not rows_to_process:
        raise HTTPException(status_code=400, detail="No data rows found in the uploaded file.")

    imported_items = []
    errors = []
    skipped_count = 0

    def get_val(row, *aliases):
        for a in aliases:
            for k in row:
                k_clean = k.replace(" ", "_").replace("-", "_").lower()
                if k_clean == a or a in k_clean:
                    v = row[k]
                    if v is not None:
                        return v
        return None

    for line_num, row in enumerate(rows_to_process, start=2):
        med_name = get_val(row, "medicine_name", "product_name", "medicine", "name", "item_name")
        if not med_name or str(med_name).strip() == "":
            skipped_count += 1
            continue

        med_name_str = str(med_name).strip()
        if med_name_str.lower() in ["medicine_name", "product_name", "column name"]:
            skipped_count += 1
            continue

        batch_no = get_val(row, "batch_no", "batch_number", "batch", "lot_no", "lot")
        if not batch_no or str(batch_no).strip() == "":
            batch_no = f"IMP-{int(datetime.utcnow().timestamp())}"
        batch_no_str = str(batch_no).strip()

        raw_exp = get_val(row, "expiry_date", "exp_date", "expiry", "exp")
        exp_date = parse_date(raw_exp)
        if not exp_date:
            errors.append(f"Row {line_num} ('{med_name_str}'): Invalid expiry date '{raw_exp}'. Use DD-MM-YYYY.")
            continue

        raw_mfd = get_val(row, "mfd_date", "mfg_date", "manufacturing_date", "mfd")
        mfd_date = parse_date(raw_mfd) if raw_mfd else None

        raw_qty = get_val(row, "quantity", "qty", "stock_qty", "packs")
        try:
            qty = int(float(str(raw_qty).replace(",", "").strip())) if raw_qty is not None else 1
            if qty <= 0:
                qty = 1
        except Exception:
            qty = 1

        raw_mrp = get_val(row, "mrp", "unit_price", "selling_price", "price")
        raw_purchase = get_val(row, "purchase_price", "purchase_rate", "cost_price", "cost", "rate")

        try:
            mrp_val = float(str(raw_mrp).replace("₹", "").replace(",", "").strip()) if raw_mrp is not None else 0.0
        except Exception:
            mrp_val = 0.0

        try:
            purchase_val = float(str(raw_purchase).replace("₹", "").replace(",", "").strip()) if raw_purchase is not None else round(mrp_val * 0.7, 2)
        except Exception:
            purchase_val = round(mrp_val * 0.7, 2)

        raw_pack = get_val(row, "pack_size_label", "pack_size", "packaging", "pack", "units_per_pack")
        pack_label = str(raw_pack).strip() if raw_pack else "1x10 Tablets"
        units_per_pack = parse_pack_units(pack_label)

        raw_gst = get_val(row, "gst_percent", "gst_rate", "gst", "tax_percent")
        try:
            gst_val = float(str(raw_gst).replace("%", "").strip()) if raw_gst is not None else 12.0
        except Exception:
            gst_val = 12.0

        brand_val = str(get_val(row, "manufacturer", "brand", "company") or "").strip() or None
        hsn_val = str(get_val(row, "hsn_code", "hsn") or "3004").strip()
        rack_val = str(get_val(row, "rack_location", "location", "rack", "shelf") or "").strip() or None

        try:
            add_req = schemas.InventoryAddRequest(
                product_name=med_name_str,
                brand=brand_val,
                category="allopathy",
                hsn_code=hsn_val,
                gst_rate=gst_val,
                batch_number=batch_no_str,
                quantity=qty,
                purchase_price=purchase_val,
                unit_price=mrp_val,
                units_per_pack=units_per_pack,
                expiry_date=exp_date.strftime("%Y-%m-%d"),
                manufacturing_date=mfd_date.strftime("%Y-%m-%d") if mfd_date else None,
                pack_size_label=pack_label,
                location=rack_val,
                duplicate_mode="merge"
            )
            prod = add_real_inventory_item(db=db, data=add_req, user_id=user_id, do_commit=False)
            imported_items.append({
                "id": prod.id,
                "product_name": prod.product_name,
                "batch_number": prod.batch_number,
                "quantity": prod.quantity,
                "mrp": prod.unit_price,
                "expiry_date": str(prod.expiry_date)
            })
        except Exception as add_err:
            errors.append(f"Row {line_num} ('{med_name_str}'): {str(add_err)}")

    if imported_items:
        db.commit()

    return {
        "success": len(imported_items) > 0 or len(errors) == 0,
        "total_rows": len(rows_to_process),
        "imported_count": len(imported_items),
        "skipped_count": skipped_count,
        "errors": errors,
        "imported_items": imported_items[:20],
        "message": f"Successfully imported {len(imported_items)} medicine batches into active inventory."
        if imported_items else "No items imported."
    }


# ==========================================
# RESTOCK SUGGESTIONS & INVENTORY REORDER ENGINE
# ==========================================
# RESTOCK SUGGESTIONS & INVENTORY REORDER ENGINE
# ==========================================

_USER_RESTOCK_DATA_CACHE: Dict[int, Tuple[list, dict, dict, float]] = {}
_RESTOCK_CACHE = {}

def invalidate_restock_cache(user_id: Optional[int] = None):
    """Invalidates cached restock calculation for a user or globally."""
    global _RESTOCK_CACHE, _USER_RESTOCK_DATA_CACHE
    if user_id is not None:
        _USER_RESTOCK_DATA_CACHE.pop(user_id, None)
        keys_to_del = [k for k in list(_RESTOCK_CACHE.keys()) if k[0] == user_id]
        for k in keys_to_del:
            _RESTOCK_CACHE.pop(k, None)
    else:
        _RESTOCK_CACHE.clear()
        _USER_RESTOCK_DATA_CACHE.clear()


def get_restock_suggestions(
    db: Session,
    user_id: int,
    multiplier: float = 3.0,
    reason_filter: Optional[str] = None,
    sort_by: str = "demand",
    search: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Computes an intelligent, demand-aware restock suggestion list:
    1. Out of stock (sellable stock == 0 and total stock == 0)
    2. Expired (all existing batches expired, zero sellable stock)
    3. Low stock relative to demand (sellable stock < 7-day demand velocity based on 30-day sales)
    
    Ranks suggestions by urgency and 30-day sales demand so fast-selling medicines appear at the top.
    """
    cache_key = (user_id, float(multiplier or 3.0), reason_filter, sort_by, (search or "").strip().lower())
    now = time.time()
    if cache_key in _RESTOCK_CACHE:
        cached_res, ts = _RESTOCK_CACHE[cache_key]
        if now - ts < 15:
            return cached_res

    import math

    today = date.today()
    cutoff_30d = datetime.utcnow() - timedelta(days=30)
    reorder_multiplier = float(multiplier if multiplier and multiplier > 0 else 3.0)

    if user_id in _USER_RESTOCK_DATA_CACHE and (now - _USER_RESTOCK_DATA_CACHE[user_id][3] < 30):
        products_data, sales_map, sales_by_id = _USER_RESTOCK_DATA_CACHE[user_id][:3]
    else:
        # 1. Fetch all non-deleted products for this shop (select only necessary columns)
        products = (
            db.query(
                models.Product.id,
                models.Product.product_name,
                models.Product.brand,
                models.Product.category,
                models.Product.unit_price,
                models.Product.units_per_pack,
                models.Product.tablets_per_strip,
                models.Product.quantity,
                models.Product.expiry_date,
                models.Product.batch_number,
            )
            .filter(
                models.Product.user_id == user_id,
                models.Product.is_deleted == False,
            )
            .all()
        )
        products_data = [
            {
                "id": p.id,
                "product_name": p.product_name,
                "brand": p.brand,
                "category": p.category,
                "composition": None,
                "pack_size_label": None,
                "unit_price": p.unit_price or 0.0,
                "units_per_pack": p.units_per_pack or p.tablets_per_strip or 10,
                "quantity": p.quantity or 0,
                "expiry_date": p.expiry_date,
                "batch_number": p.batch_number,
            }
            for p in products
        ]

        # 2. Fetch sales demand aggregated over last 30 days for this shop
        sales_30d = (
            db.query(
                models.SaleItem.product_id,
                func.lower(func.trim(models.SaleItem.product_name)).label("norm_name"),
                func.sum(models.SaleItem.quantity).label("total_sold_qty"),
                func.count(models.SaleItem.id).label("bill_count"),
            )
            .join(models.Sale, models.SaleItem.sale_id == models.Sale.id)
            .filter(
                models.Sale.user_id == user_id,
                models.Sale.created_at >= cutoff_30d,
                or_(models.Sale.return_status != "returned", models.Sale.return_status.is_(None)),
            )
            .group_by(models.SaleItem.product_id, func.lower(func.trim(models.SaleItem.product_name)))
            .all()
        )

        sales_map = {}
        sales_by_id = {}
        for row in sales_30d:
            qty = float(row.total_sold_qty or 0)
            bills = int(row.bill_count or 0)
            if row.norm_name:
                if row.norm_name not in sales_map:
                    sales_map[row.norm_name] = {"total_sold_qty": 0.0, "bill_count": 0}
                sales_map[row.norm_name]["total_sold_qty"] += qty
                sales_map[row.norm_name]["bill_count"] += bills
            if row.product_id:
                if row.product_id not in sales_by_id:
                    sales_by_id[row.product_id] = {"total_sold_qty": 0.0, "bill_count": 0}
                sales_by_id[row.product_id]["total_sold_qty"] += qty
                sales_by_id[row.product_id]["bill_count"] += bills

        _USER_RESTOCK_DATA_CACHE[user_id] = (products_data, sales_map, sales_by_id, now)

    # 3. Group inventory by normalized medicine name directly from products and sales
    med_groups: Dict[str, Dict[str, Any]] = {}

    for p in products_data:
        n_name = (p["product_name"] or "").strip().lower()
        if not n_name:
            continue
        if n_name not in med_groups:
            med_groups[n_name] = {
                "id": p["id"],
                "product_name": p["product_name"].strip(),
                "brand": p["brand"] or "",
                "category": p["category"] or "allopathy",
                "composition": p["composition"],
                "pack_size_label": p["pack_size_label"],
                "unit_price": p["unit_price"] or 0.0,
                "units_per_pack": p["units_per_pack"] or 10,
                "total_stock": 0,
                "sellable_stock": 0,
                "expired_stock": 0,
                "batches": [],
                "nearest_expiry": None,
            }

        g = med_groups[n_name]
        qty = p["quantity"] or 0
        g["total_stock"] += qty

        is_expired = p["expiry_date"] < today if p["expiry_date"] else False
        if is_expired:
            g["expired_stock"] += qty
        else:
            g["sellable_stock"] += qty
            if g["nearest_expiry"] is None or p["expiry_date"] < g["nearest_expiry"]:
                g["nearest_expiry"] = p["expiry_date"]

        g["batches"].append({
            "batch_number": p["batch_number"] or "N/A",
            "quantity": qty,
            "expiry_date": str(p["expiry_date"]) if p["expiry_date"] else "N/A",
            "is_expired": is_expired,
        })

        if p["unit_price"] and p["unit_price"] > 0:
            g["unit_price"] = p["unit_price"]
        if p["brand"] and not g["brand"]:
            g["brand"] = p["brand"]

    # Include medicines sold in last 30 days that have 0 current product stock records
    for norm_name, s_data in sales_map.items():
        if norm_name not in med_groups:
            med_groups[norm_name] = {
                "id": None,
                "product_name": norm_name.title(),
                "brand": "",
                "category": "allopathy",
                "composition": None,
                "pack_size_label": None,
                "unit_price": 0.0,
                "units_per_pack": 10,
                "total_stock": 0,
                "sellable_stock": 0,
                "expired_stock": 0,
                "batches": [],
                "nearest_expiry": None,
            }

    # 5. Evaluate restock status & suggested reorder quantities
    suggestions = []
    out_of_stock_count = 0
    expired_count = 0
    low_stock_count = 0
    total_reorder_units = 0
    estimated_reorder_value = 0.0

    total_30d_sales_units = 0.0
    total_30d_bill_count = 0

    for norm_name, g in med_groups.items():
        s_data = sales_map.get(norm_name)
        if not s_data and g.get("id") and g["id"] in sales_by_id:
            s_data = sales_by_id[g["id"]]
        if not s_data:
            s_data = {"total_sold_qty": 0.0, "bill_count": 0}

        sold_30d = float(s_data["total_sold_qty"])
        bill_count_30d = int(s_data["bill_count"])
        total_30d_sales_units += sold_30d
        total_30d_bill_count += bill_count_30d

        avg_daily_sales = round(sold_30d / 30.0, 2)
        avg_weekly_sales = round(sold_30d * (7.0 / 30.0), 2)

        sellable_stock = g["sellable_stock"]
        expired_stock = g["expired_stock"]
        total_stock = g["total_stock"]

        reason = None
        reason_label = None
        urgency_level = "Moderate"
        urgency_weight = 0
        days_of_stock = None

        if sellable_stock == 0:
            if expired_stock > 0:
                reason = "EXPIRED"
                reason_label = "Expired (Zero Sellable Stock)"
                urgency_level = "Critical" if sold_30d > 0 else "High"
                urgency_weight = 2
                days_of_stock = 0.0
                expired_count += 1
            else:
                reason = "OUT_OF_STOCK"
                reason_label = "Out of Stock"
                urgency_level = "Critical" if sold_30d > 0 else "High"
                urgency_weight = 3
                days_of_stock = 0.0
                out_of_stock_count += 1
        elif sellable_stock > 0:
            if avg_daily_sales > 0:
                days_of_stock = round(sellable_stock / avg_daily_sales, 1)
                if days_of_stock < 7.0 or sellable_stock <= 3:
                    reason = "LOW_STOCK"
                    reason_label = f"Low Stock ({days_of_stock}d stock left)"
                    urgency_level = "High" if days_of_stock <= 3.0 else "Moderate"
                    urgency_weight = 1
                    low_stock_count += 1
            else:
                # No recent sales in 30d, but stock is down to minimal (<= 2)
                if sellable_stock <= 2:
                    reason = "LOW_STOCK"
                    reason_label = "Low Stock (Safety Threshold)"
                    urgency_level = "Moderate"
                    urgency_weight = 0
                    days_of_stock = None
                    low_stock_count += 1

        if not reason:
            continue

        # Compute suggested reorder quantity: (avg weekly sales) * multiplier
        if avg_weekly_sales > 0:
            raw_reorder = avg_weekly_sales * reorder_multiplier
            suggested_reorder_qty = max(int(math.ceil(raw_reorder)), 5)
        else:
            # Never sold or 0 recent velocity: suggest default reorder pack
            suggested_reorder_qty = 10

        est_cost = round(suggested_reorder_qty * (g["unit_price"] or 0.0), 2)
        total_reorder_units += suggested_reorder_qty
        estimated_reorder_value += est_cost

        urgency_score = (urgency_weight * 1000) + (sold_30d * 25) + (100 if sellable_stock == 0 else max(0, 50 - sellable_stock))

        item_dict = {
            "id": g["id"],
            "product_name": g["product_name"],
            "brand": g["brand"] or "Generic",
            "category": g["category"] or "allopathy",
            "composition": g["composition"],
            "pack_size_label": g["pack_size_label"],
            "unit_price": g["unit_price"],
            "units_per_pack": g["units_per_pack"],
            "sellable_stock": sellable_stock,
            "expired_stock": expired_stock,
            "total_stock": total_stock,
            "nearest_expiry": str(g["nearest_expiry"]) if g["nearest_expiry"] else None,
            "sales_30d": sold_30d,
            "bill_count_30d": bill_count_30d,
            "avg_daily_sales": avg_daily_sales,
            "avg_weekly_sales": avg_weekly_sales,
            "days_of_stock_remaining": days_of_stock,
            "reason": reason,
            "reason_label": reason_label,
            "urgency_level": urgency_level,
            "urgency_score": round(urgency_score, 2),
            "suggested_reorder_qty": suggested_reorder_qty,
            "estimated_reorder_cost": est_cost,
            "batches": g["batches"],
        }
        suggestions.append(item_dict)

    # Filter by Reason if specified
    if reason_filter and reason_filter.lower() != "all":
        rf = reason_filter.lower()
        if rf in ["out_of_stock", "outofstock"]:
            suggestions = [s for s in suggestions if s["reason"] == "OUT_OF_STOCK"]
        elif rf in ["expired"]:
            suggestions = [s for s in suggestions if s["reason"] == "EXPIRED"]
        elif rf in ["low_stock", "lowstock"]:
            suggestions = [s for s in suggestions if s["reason"] == "LOW_STOCK"]

    # Filter by Search term
    if search and search.strip():
        q_term = search.strip().lower()
        suggestions = [
            s for s in suggestions
            if q_term in s["product_name"].lower()
            or (s["brand"] and q_term in s["brand"].lower())
            or (s["composition"] and q_term in s["composition"].lower())
        ]

    # Sort options
    if sort_by == "name":
        suggestions.sort(key=lambda s: s["product_name"].lower())
    elif sort_by == "stock":
        suggestions.sort(key=lambda s: (s["sellable_stock"], -s["sales_30d"]))
    elif sort_by == "sales":
        suggestions.sort(key=lambda s: s["sales_30d"], reverse=True)
    else:  # "demand" / default urgency ranking
        suggestions.sort(key=lambda s: s["urgency_score"], reverse=True)

    has_sales_history = total_30d_sales_units > 0 or len(sales_map) > 0

    final_output = {
        "success": True,
        "summary": {
            "total_suggestions": len(suggestions),
            "out_of_stock_count": out_of_stock_count,
            "expired_count": expired_count,
            "low_stock_count": low_stock_count,
            "total_reorder_units": total_reorder_units,
            "estimated_reorder_value": round(estimated_reorder_value, 2),
            "multiplier": reorder_multiplier,
            "total_products_evaluated": len(products_data),
            "has_sales_history": has_sales_history,
            "total_30d_sales_units": round(total_30d_sales_units, 1),
            "total_30d_bill_count": total_30d_bill_count,
        },
        "suggestions": suggestions,
    }
    _RESTOCK_CACHE[cache_key] = (final_output, now)
    return final_output


# ==========================================
# ERP PRIORITY 1 CRUD FUNCTIONS
# ==========================================

import uuid
from datetime import datetime

def create_purchase_invoice(db: Session, obj_in: schemas.PurchaseInvoiceCreate, user_id: int):
    """Processes purchase invoice and automatically updates product inventory."""
    # Check if supplier exists
    supplier = db.query(models.Supplier).filter(
        models.Supplier.id == obj_in.supplier_id,
        models.Supplier.user_id == user_id
    ).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found or access denied.")

    # Create PurchaseInvoice
    db_invoice = models.PurchaseInvoice(
        user_id=user_id,
        supplier_id=obj_in.supplier_id,
        invoice_number=obj_in.invoice_number,
        invoice_date=obj_in.invoice_date,
        total_amount=obj_in.total_amount,
        tax_amount=obj_in.tax_amount,
        payment_status=obj_in.payment_status
    )
    db.add(db_invoice)
    db.flush()  # Generate db_invoice.id

    for item in obj_in.items:
        # Check if product with ID exists in inventory
        db_product = db.query(models.Product).filter(
            models.Product.id == item.product_id,
            models.Product.user_id == user_id
        ).first()

        # If matching product found, increment quantity and update price
        if db_product:
            db_product.quantity += item.quantity
            db_product.purchase_price = item.purchase_price
            db_product.unit_price = item.mrp
            db_product.gst_rate = item.gst_rate
            db_product.expiry_date = item.expiry_date
        else:
            raise HTTPException(status_code=404, detail=f"Product with ID {item.product_id} not found.")

        # Create PurchaseItem record linked to invoice
        db_item = models.PurchaseItem(
            purchase_invoice_id=db_invoice.id,
            product_id=db_product.id,
            batch_number=item.batch_number,
            quantity=item.quantity,
            purchase_price=item.purchase_price,
            mrp=item.mrp,
            gst_rate=item.gst_rate,
            expiry_date=item.expiry_date
        )
        db.add(db_item)

        # Log Inventory Transaction
        db_txn = models.InventoryTransaction(
            transaction_id=f"TXN-{datetime.utcnow():%Y%m%d%H%M%S}-{uuid.uuid4().hex[:8].upper()}",
            shop_id=user_id,
            product_id=db_product.id,
            transaction_type="purchase",
            quantity=item.quantity,
            unit_price=item.mrp,
            purchase_price=item.purchase_price,
            total_price=round(item.quantity * item.purchase_price, 2),
            final_price=round(item.quantity * item.purchase_price, 2)
        )
        db.add(db_txn)

    db.commit()
    db.refresh(db_invoice)
    return db_invoice


def get_purchase_invoices(db: Session, user_id: int, skip: int = 0, limit: int = 50):
    return db.query(models.PurchaseInvoice).filter(
        models.PurchaseInvoice.user_id == user_id
    ).offset(skip).limit(limit).all()


def get_purchase_invoice(db: Session, purchase_id: int, user_id: int):
    return db.query(models.PurchaseInvoice).filter(
        models.PurchaseInvoice.id == purchase_id,
        models.PurchaseInvoice.user_id == user_id
    ).first()
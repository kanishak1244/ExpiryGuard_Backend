import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule


def generate_inventory_import_template() -> io.BytesIO:
    """
    Generates a polished, professional-grade .xlsx inventory import template
    matching Zoho Inventory, Vyapar, and Tally quality standards.
    Structured into two sheets: 'Instructions' (opens by default) and 'Data'.
    """
    wb = openpyxl.Workbook()

    # Define color palette & typography
    FONT_FAMILY = "Calibri"
    BRAND_TEAL = "0F766E"
    BRAND_DARK_NAVY = "1E293B"
    SLATE_GRAY = "64748B"
    SLATE_LIGHT_BG = "F8FAFC"
    BORDER_COLOR = "CBD5E1"
    WARNING_BG = "FEF2F2"
    WARNING_BORDER = "FCA5A5"
    WARNING_TEXT = "991B1B"
    REQUIRED_GREEN = "047857"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    header_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="medium", color=BRAND_TEAL),
    )

    # =========================================================================
    # SHEET 1: Instructions (First sheet, active by default)
    # =========================================================================
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst.views.sheetView[0].showGridLines = True

    # 1. Title Banner (Merged A1:D2)
    ws_inst.merge_cells("A1:D2")
    title_cell = ws_inst["A1"]
    title_cell.value = "ExpiryGuard -- Inventory Bulk Import Template"
    title_cell.font = Font(name=FONT_FAMILY, size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=BRAND_TEAL, end_color=BRAND_TEAL, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_inst.row_dimensions[1].height = 24
    ws_inst.row_dimensions[2].height = 24

    # 2. Intro Text
    ws_inst["A4"].value = "Welcome to the ExpiryGuard bulk inventory upload guide."
    ws_inst["A4"].font = Font(name=FONT_FAMILY, size=11, bold=True, color=BRAND_DARK_NAVY)

    ws_inst["A5"].value = "Follow the column specifications below to populate the 'Data' sheet. Save your completed file and upload it through the Add Inventory -> Bulk Import modal."
    ws_inst["A5"].font = Font(name=FONT_FAMILY, size=10.5, color="475569")
    ws_inst.row_dimensions[4].height = 18
    ws_inst.row_dimensions[5].height = 18

    # 3. Instruction Table Headers
    inst_headers = ["Column Name", "Required?", "Format / Example", "Field Guidance & Notes"]
    ws_inst.row_dimensions[7].height = 26

    for col_idx, h_text in enumerate(inst_headers, start=1):
        cell = ws_inst.cell(row=7, column=col_idx, value=h_text)
        cell.font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BRAND_DARK_NAVY, end_color=BRAND_DARK_NAVY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_idx <= 2 else "left", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 4. Instruction Table Rows
    column_specs = [
        ("medicine_name", "Yes", "Dolo 650mg Tablet", "Brand or generic medicine name as sold to customers (must not be empty)."),
        ("manufacturer", "No", "Micro Labs Ltd", "Pharmaceutical manufacturer, brand owner, or marketing company."),
        ("hsn_code", "No", "3004", "4-digit or 6-digit Indian GST HSN code (standard pharma baseline: 3004)."),
        ("batch_no", "Yes", "DL7820", "Unique Batch/Lot number printed on the medicine box or foil strip."),
        ("mfd_date", "No", "15-01-2026", "Manufacturing date formatted as DD-MM-YYYY (e.g. 15-01-2026)."),
        ("expiry_date", "Yes", "30-11-2027", "Expiry date formatted as DD-MM-YYYY (must be a valid future date)."),
        ("pack_size_label", "Yes", "10x10 or 15 Tablets", "Packaging unit label (e.g. '10x10', '15 Tablets', '100ml Bottle')."),
        ("quantity", "Yes", "50", "Total sealed packs / strips available in stock (whole number > 0)."),
        ("purchase_price", "Yes", "18.50", "Supplier net purchase rate per pack in INR (before/including input tax)."),
        ("mrp", "Yes", "33.60", "Maximum Retail Price (MRP) printed on packaging per pack in INR."),
        ("gst_percent", "No", "12", "Applicable Indian GST slab: 0, 5, 12, 18, or 28 (defaults to 12% if blank)."),
        ("rack_location", "No", "Rack-A2 / Shelf-3", "Pharmacy storage cabinet, shelf, drawer, or rack location identifier."),
    ]

    for idx, (col_name, req, example, notes) in enumerate(column_specs, start=8):
        ws_inst.row_dimensions[idx].height = 22
        bg_fill = PatternFill(start_color=SLATE_LIGHT_BG if idx % 2 == 0 else "FFFFFF", end_color=SLATE_LIGHT_BG if idx % 2 == 0 else "FFFFFF", fill_type="solid")

        c1 = ws_inst.cell(row=idx, column=1, value=col_name)
        c1.font = Font(name=FONT_FAMILY, size=10.5, bold=True, color="0F172A")
        c1.fill = bg_fill
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c1.border = thin_border

        c2 = ws_inst.cell(row=idx, column=2, value=req)
        c2.font = Font(name=FONT_FAMILY, size=10.5, bold=(req == "Yes"), color=REQUIRED_GREEN if req == "Yes" else SLATE_GRAY)
        c2.fill = bg_fill
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = thin_border

        c3 = ws_inst.cell(row=idx, column=3, value=example)
        c3.font = Font(name=FONT_FAMILY, size=10, italic=True, color="334155")
        c3.fill = bg_fill
        c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c3.border = thin_border

        c4 = ws_inst.cell(row=idx, column=4, value=notes)
        c4.font = Font(name=FONT_FAMILY, size=10, color="475569")
        c4.fill = bg_fill
        c4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c4.border = thin_border

    # 5. Common Mistakes to Avoid Section
    mistakes_start = len(column_specs) + 10
    ws_inst.merge_cells(start_row=mistakes_start, start_column=1, end_row=mistakes_start, end_column=4)
    warn_header = ws_inst.cell(row=mistakes_start, column=1, value="Common Mistakes to Avoid & Best Practices")
    warn_header.font = Font(name=FONT_FAMILY, size=11, bold=True, color=WARNING_TEXT)
    warn_header.fill = PatternFill(start_color=WARNING_BG, end_color=WARNING_BG, fill_type="solid")
    warn_header.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_inst.row_dimensions[mistakes_start].height = 24

    mistake_items = [
        "1. Do NOT rename, remove, or reorder column headers in the 'Data' sheet.",
        "2. Ensure dates are written strictly in DD-MM-YYYY format (e.g. 31-12-2027, NOT 12/31/2027).",
        "3. Do not leave required columns blank (medicine_name, batch_no, expiry_date, pack_size_label, quantity, purchase_price, mrp).",
        "4. Sample rows in the 'Data' sheet are examples for your reference -- overwrite or replace them with your actual data.",
        "5. GST percent must be a standard slab number: 0, 5, 12, 18, or 28.",
    ]

    for m_idx, text in enumerate(mistake_items, start=mistakes_start + 1):
        ws_inst.merge_cells(start_row=m_idx, start_column=1, end_row=m_idx, end_column=4)
        c = ws_inst.cell(row=m_idx, column=1, value=text)
        c.font = Font(name=FONT_FAMILY, size=10, color="7F1D1D")
        c.fill = PatternFill(start_color=WARNING_BG, end_color=WARNING_BG, fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_inst.row_dimensions[m_idx].height = 20

    # Auto-adjust column widths for Instructions sheet
    inst_col_widths = {1: 22, 2: 14, 3: 30, 4: 75}
    for col_idx, width in inst_col_widths.items():
        ws_inst.column_dimensions[get_column_letter(col_idx)].width = width

    # Lock / Protect Instructions Sheet
    ws_inst.protection.sheet = True
    ws_inst.protection.enable()

    # =========================================================================
    # SHEET 2: Data (Where the shopkeeper fills entries)
    # =========================================================================
    ws_data = wb.create_sheet(title="Data")
    ws_data.views.sheetView[0].showGridLines = True

    data_columns = [
        ("medicine_name", 32),
        ("manufacturer", 24),
        ("hsn_code", 14),
        ("batch_no", 18),
        ("mfd_date", 16),
        ("expiry_date", 16),
        ("pack_size_label", 18),
        ("quantity", 14),
        ("purchase_price", 18),
        ("mrp", 16),
        ("gst_percent", 14),
        ("rack_location", 18),
    ]

    # Header Row (Row 1)
    ws_data.row_dimensions[1].height = 28
    for col_idx, (col_name, width) in enumerate(data_columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BRAND_TEAL, end_color=BRAND_TEAL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
        cell.protection = Protection(locked=True)
        ws_data.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze Header Row
    ws_data.freeze_panes = "A2"

    # Pre-filled 3 Example / Sample Rows (Rows 2 to 4)
    sample_rows = [
        ["Augmentin 625 Duo Tablet", "GlaxoSmithKline", "3004", "AUG8921", "10-01-2026", "31-08-2027", "1x10 Tablets", 40, 142.50, 204.80, 12, "Rack-A1"],
        ["Pan 40 Tablet", "Alkem Laboratories", "3004", "PN4401", "05-02-2026", "30-01-2028", "1x15 Tablets", 100, 88.00, 155.00, 12, "Rack-B3"],
        ["Azee 500mg Tablet", "Cipla Ltd", "3004", "AZ5520", "12-12-2025", "30-11-2027", "1x5 Tablets", 60, 71.20, 119.50, 12, "Rack-C2"],
    ]

    sample_font = Font(name=FONT_FAMILY, size=10.5, italic=True, color="64748B")
    sample_fill = PatternFill(start_color=SLATE_LIGHT_BG, end_color=SLATE_LIGHT_BG, fill_type="solid")

    for row_idx, row_data in enumerate(sample_rows, start=2):
        ws_data.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
            cell.font = sample_font
            cell.fill = sample_fill
            cell.border = thin_border
            cell.protection = Protection(locked=False)

            if col_idx in [5, 6]:  # mfd_date, expiry_date
                cell.number_format = "DD-MM-YYYY"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8:  # quantity
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [9, 10]:  # purchase_price, mrp
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 11:  # gst_percent
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [3, 4]:  # hsn_code, batch_no
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Unlock Data Entry Cells (Rows 5 to 500) so shopkeepers can enter data freely
    regular_font = Font(name=FONT_FAMILY, size=11, color="0F172A")
    for r in range(5, 501):
        ws_data.row_dimensions[r].height = 21
        for c in range(1, 13):
            cell = ws_data.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border
            cell.protection = Protection(locked=False)

            if c in [5, 6]:
                cell.number_format = "DD-MM-YYYY"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 8:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c in [9, 10]:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c in [3, 4, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data Validation Dropdowns
    # 1. GST Percent Dropdown Validation (0, 5, 12, 18, 28)
    gst_dv = DataValidation(
        type="list",
        formula1='"0,5,12,18,28"',
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    gst_dv.error = "Please select a valid GST slab: 0, 5, 12, 18, or 28%."
    gst_dv.errorTitle = "Invalid GST Slab"
    gst_dv.prompt = "Select official GST rate (0%, 5%, 12%, 18%, 28%)."
    gst_dv.promptTitle = "GST Slab Selection"
    ws_data.add_data_validation(gst_dv)
    gst_dv.add("K2:K500")

    # 2. Quantity Validation (> 0)
    qty_dv = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1=0,
        allow_blank=False,
        showErrorMessage=True,
    )
    qty_dv.error = "Quantity must be a positive whole number greater than 0."
    qty_dv.errorTitle = "Invalid Quantity"
    ws_data.add_data_validation(qty_dv)
    qty_dv.add("H2:H500")

    # Conditional Formatting: Highlight expiry_date cells in light red if date < TODAY()
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_font = Font(name=FONT_FAMILY, size=10.5, color="991B1B", bold=True)
    ws_data.conditional_formatting.add(
        "F2:F500",
        CellIsRule(operator="lessThan", formula=["TODAY()"], fill=red_fill, font=red_font)
    )

    # Protect Data sheet (Lock header row while allowing unlocked data cells to be edited)
    ws_data.protection.sheet = True
    ws_data.protection.selectLockedCells = True
    ws_data.protection.selectUnlockedCells = True
    ws_data.protection.formatCells = True
    ws_data.protection.enable()

    # Ensure Instructions sheet is the active sheet upon opening
    wb.active = 0

    # Save to BytesIO stream
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream
